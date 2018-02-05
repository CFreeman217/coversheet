
import os
import csv
import re
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter

part1_regex = re.compile(r"""(                    
    (\w*\d_)       # Lab Number
    (\s)?               # Space
    (\d*)               # Number of Samples
    (\s)                # Space
    (\d*)             # Sampling Freq
    (hz)
    (\.[a-zA-Z]{2,4})   # Suffix
                        )""", re.VERBOSE)

part2_regex = re.compile(r"""(                    
    (\w{3}\d{1}_)       # Lab Number
    (FFT\s|p2_)           # Part number or test
    (raw\sdata|resample)?
    (\s)?
    (\d*)               # Number of Samples
    (_)?
    (\s)?           # Space
    (\d*|\w{2,4})             # Sampling Freq
    
    (\.[a-zA-Z]{2,4})   # Suffix
                        )""", re.VERBOSE)

def readFile(fileName):

    with open(fileName) as file:
        reader = csv.reader(file, delimiter='\t')
        next(reader, None)
        data = list(reader)
    # print(data)
    return data

def collect_data():
    
    collection = []
    for dirpath, dirnames, files in os.walk('c:\\Users\\freem\\Documents\\Python'):
        for name in files:
            
            if name.lower().startswith('lab3') and name.lower().endswith('.txt'):
                os.chdir(dirpath)
                print(name)
                collection.append(name)
                collection.append(readFile(name))
                os.chdir('..')
    return collection

def write_data(input_data):
    wb = Workbook()
    for list_item in input_data:
        # print(list_item)
        if len(list_item) == 1 and list_item.startswith('Lab3'):
            wb.create_sheet(title=list_item)
        else:
            for datapiece in list_item:
                wb.active.append(datapiece[0])

    wb.save(filename = "COMBINE")
    
data = collect_data()
write_data(data)